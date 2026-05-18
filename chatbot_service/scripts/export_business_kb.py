"""Export Playfunia business content from Supabase into a structured Excel KB.

Goal: produce data that supports BOTH semantic (vector) and structured (SQL) queries.

Output: chatbot_service/data/business_knowledge_base.xlsx
Sheets:
  knowledge_base   — unified doc-per-row: doc_id, category, slug, name, summary,
                     tags, source_table, entity_id, source_updated_at
                     (`summary` is the text we embed; one self-contained paragraph)
  facts            — Q&A pairs: doc_id, category, question, answer
                     (quotable, deterministic snippets the bot can return verbatim)
  locations        — name, street, city, state, postal_code, phone, email
  hours            — location, day_of_week, opens, closes, is_closed
  tickets          — name, price_usd, child_count, requires_waiver,
                     requires_grip_socks, description
  memberships      — name, monthly_price_usd, max_children, max_adults,
                     visits_per_month, guest_passes_per_month,
                     member_discount_percent, benefits
  party_packages   — name, price_usd, base_children, base_room_hours,
                     includes_food, includes_drinks, includes_decor,
                     extra_child_price_usd, extra_adult_price_usd,
                     cleaning_fee_usd, features, terms
  party_addons     — code, label, price_usd, price_type, description
  events           — title, start_date, end_date, description
  promotions       — kind, label_or_code, scope, discount, starts_at, ends_at,
                     description
  pricing_fees     — key, value_usd_or_percent, unit, description, audience
  jobs             — title, department, employment_type, location, pay_range,
                     minimum_age, schedule_notes, summary
"""
from __future__ import annotations

import json
import os
import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(ROOT / ".env")

PG = dict(
    host="db.wzmcmbkouodsfbfxaozd.supabase.co",
    port=5432,
    user="postgres",
    dbname="postgres",
    password=os.environ["supabase_db_password"],
)

OUT_PATH = ROOT / "chatbot_service" / "data" / "business_knowledge_base.xlsx"

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# ---------- helpers --------------------------------------------------------------

def jsonable(v: Any) -> Any:
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (list, tuple)):
        return [jsonable(x) for x in v]
    if isinstance(v, dict):
        return {k: jsonable(x) for k, x in v.items()}
    return v


def money(v) -> str:
    if v is None:
        return "—"
    f = float(v)
    return f"${f:,.0f}" if f.is_integer() else f"${f:,.2f}"


def fmt_time(v) -> str:
    if v is None:
        return ""
    if isinstance(v, time):
        s = v.strftime("%I:%M %p").lstrip("0")
        return s.replace(":00 ", " ")  # 10 AM instead of 10:00 AM
    return str(v)


def fmt_date(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%b %-d, %Y")
    return str(v)


def slug(s: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")
    return s or "item"


def fetch_all(cur, table: str, where: str = "", order: str = "") -> list[dict]:
    cur.execute(f"SELECT * FROM public.{table} {where} {order}".strip())
    return [dict(r) for r in cur.fetchall()]


# ---------- workbook scaffolding ------------------------------------------------

class KB:
    """Accumulator: structured sheets + unified doc list + facts."""

    def __init__(self):
        self.docs: list[dict] = []   # rows for knowledge_base
        self.facts: list[dict] = []  # rows for facts
        self.sheets: dict[str, dict] = {}  # name -> {"cols":[...], "rows":[[...]]}

    def add_sheet(self, name: str, cols: list[str]):
        self.sheets[name] = {"cols": cols, "rows": []}

    def row(self, sheet: str, **vals):
        s = self.sheets[sheet]
        s["rows"].append([vals.get(c) for c in s["cols"]])

    def add_doc(self, *, category: str, source_table: str, entity_id, name: str,
                summary: str, tags: list[str], updated_at=None):
        doc_id = len(self.docs) + 1
        self.docs.append({
            "doc_id": doc_id,
            "category": category,
            "slug": f"{category}:{slug(name)}",
            "name": name,
            "summary": summary.strip(),
            "tags": ", ".join(tags),
            "source_table": source_table,
            "entity_id": entity_id,
            "source_updated_at": jsonable(updated_at),
        })
        return doc_id

    def add_fact(self, doc_id: int, category: str, question: str, answer: str):
        self.facts.append({
            "doc_id": doc_id, "category": category,
            "question": question.strip(), "answer": answer.strip(),
        })


# ---------- builders per category -----------------------------------------------

def build_locations(kb: KB, cur):
    kb.add_sheet("locations", ["location_id", "name", "street", "city", "state",
                               "postal_code", "phone", "email"])
    rows = fetch_all(cur, "locations", "WHERE is_active IS NOT FALSE", "ORDER BY location_id")
    for r in rows:
        kb.row("locations",
               location_id=r["location_id"], name=r.get("name"),
               street=r.get("address"), city=r.get("city"), state=r.get("state"),
               postal_code=r.get("postal_code"), phone=r.get("phone"), email=r.get("email"))

        name = r.get("name") or "Playfunia"
        addr_parts = [r.get("address"), r.get("city"), r.get("state"), r.get("postal_code")]
        addr = ", ".join(p for p in addr_parts if p)
        phone = r.get("phone") or "—"
        email = r.get("email") or "—"
        summary = (
            f"Playfunia ({name}) is located at {addr}. "
            f"Reach us by phone at {phone} or email at {email}."
        )
        doc_id = kb.add_doc(
            category="location", source_table="locations", entity_id=r["location_id"],
            name=f"{name} location", summary=summary,
            tags=["location", "address", "contact", "phone", "email"],
            updated_at=r.get("updated_at"),
        )
        kb.add_fact(doc_id, "location", "Where is Playfunia located?", f"{addr}.")
        kb.add_fact(doc_id, "location", "What is Playfunia's phone number?", phone)
        kb.add_fact(doc_id, "location", "What is Playfunia's email?", email)


def build_hours(kb: KB, cur):
    kb.add_sheet("hours", ["location", "day_of_week", "opens", "closes", "is_closed"])
    rows = fetch_all(cur, "store_hours", "WHERE is_active IS NOT FALSE",
                     "ORDER BY location_name, day_of_week")
    if not rows:
        return

    by_loc: dict[str, list[dict]] = {}
    for r in rows:
        loc = r.get("location_name") or "Playfunia"
        by_loc.setdefault(loc, []).append(r)
        kb.row("hours",
               location=loc,
               day_of_week=DAYS[r["day_of_week"]] if 0 <= r["day_of_week"] < 7 else f"Day{r['day_of_week']}",
               opens=fmt_time(r.get("open_time")) if not r.get("is_closed") else "",
               closes=fmt_time(r.get("close_time")) if not r.get("is_closed") else "",
               is_closed=bool(r.get("is_closed")))

    for loc, day_rows in by_loc.items():
        day_rows.sort(key=lambda x: x["day_of_week"])
        parts = []
        for r in day_rows:
            day = DAYS[r["day_of_week"]] if 0 <= r["day_of_week"] < 7 else f"Day{r['day_of_week']}"
            if r.get("is_closed"):
                parts.append(f"{day}: closed")
            else:
                parts.append(f"{day} {fmt_time(r.get('open_time'))}–{fmt_time(r.get('close_time'))}")
        summary = f"Playfunia ({loc}) hours: " + "; ".join(parts) + "."
        doc_id = kb.add_doc(
            category="hours", source_table="store_hours", entity_id=None,
            name=f"{loc} hours of operation", summary=summary,
            tags=["hours", "open", "closed", "schedule", "when"],
        )
        kb.add_fact(doc_id, "hours", "What are Playfunia's hours?", summary)


def build_tickets(kb: KB, cur):
    kb.add_sheet("tickets", ["ticket_type_id", "name", "price_usd", "child_count",
                             "requires_waiver", "requires_grip_socks", "description"])
    rows = fetch_all(cur, "ticket_types", "WHERE is_active IS NOT FALSE",
                     "ORDER BY child_count NULLS FIRST, base_price_usd")
    for r in rows:
        kb.row("tickets",
               ticket_type_id=r["ticket_type_id"], name=r.get("name"),
               price_usd=jsonable(r.get("base_price_usd")),
               child_count=r.get("child_count"),
               requires_waiver=bool(r.get("requires_waiver")),
               requires_grip_socks=bool(r.get("requires_grip_socks")),
               description=r.get("description"))

        name = r.get("name")
        price = money(r.get("base_price_usd"))
        cc = r.get("child_count") or 0
        desc = (r.get("description") or "").strip()

        if cc == 0:
            summary = f"{name} costs {price}. {desc}".strip()
        elif cc == 1:
            summary = f"{name} is {price} and gives one child unlimited play for the day."
        else:
            summary = f"{name} is {price} and gives {cc} children unlimited play for the day."

        reqs = []
        if r.get("requires_waiver"): reqs.append("a signed waiver")
        if r.get("requires_grip_socks"): reqs.append("grip socks")
        if reqs:
            summary += " Requires " + " and ".join(reqs) + "."

        doc_id = kb.add_doc(
            category="ticket", source_table="ticket_types", entity_id=r["ticket_type_id"],
            name=name, summary=summary,
            tags=["ticket", "admission", "price", "open play", f"{cc}-children"],
            updated_at=r.get("updated_at"),
        )
        kb.add_fact(doc_id, "ticket", f"How much is {name}?", price)


def build_memberships(kb: KB, cur):
    kb.add_sheet("memberships", ["plan_id", "name", "monthly_price_usd", "max_children",
                                 "max_adults", "visits_per_month", "guest_passes_per_month",
                                 "member_discount_percent", "benefits"])
    rows = fetch_all(cur, "membership_plans", "WHERE is_active IS NOT FALSE",
                     "ORDER BY monthly_price")
    for r in rows:
        benefits = r.get("benefits") or []
        kb.row("memberships",
               plan_id=r["plan_id"], name=r.get("name"),
               monthly_price_usd=jsonable(r.get("monthly_price")),
               max_children=r.get("max_children"), max_adults=r.get("max_adults"),
               visits_per_month=r.get("visits_per_month"),
               guest_passes_per_month=r.get("guest_passes_per_month") or 0,
               member_discount_percent=jsonable(r.get("discount_percent")) or 0,
               benefits="; ".join(benefits))

        name = r.get("name")
        price = money(r.get("monthly_price"))
        kids = r.get("max_children") or 0
        adults = r.get("max_adults") or 0
        visits = r.get("visits_per_month")
        visits_phrase = "unlimited visits" if visits is None else f"{visits} visits/month"
        people = f"{kids} child{'ren' if kids != 1 else ''} + {adults} adult{'s' if adults != 1 else ''}"
        guests = r.get("guest_passes_per_month") or 0
        disc = float(r.get("discount_percent") or 0)
        extras = []
        if guests:
            extras.append(f"{guests} guest pass{'es' if guests != 1 else ''} per month")
        if disc:
            extras.append(f"{disc:g}% member discount")
        extras_phrase = (" Includes " + ", ".join(extras) + ".") if extras else ""

        summary = (
            f"The {name} membership is {price}/month and covers {people} with {visits_phrase}."
            f"{extras_phrase}"
        )
        if benefits:
            summary += " Benefits: " + "; ".join(benefits) + "."

        doc_id = kb.add_doc(
            category="membership", source_table="membership_plans", entity_id=r["plan_id"],
            name=name, summary=summary,
            tags=["membership", "monthly", "plan", f"{kids}-children"],
            updated_at=r.get("updated_at"),
        )
        kb.add_fact(doc_id, "membership", f"How much is the {name}?", f"{price}/month")
        kb.add_fact(doc_id, "membership", f"Who does the {name} cover?", people)


def build_party_packages(kb: KB, cur):
    kb.add_sheet("party_packages", ["package_id", "name", "price_usd", "base_children",
                                    "base_room_hours", "includes_food", "includes_drinks",
                                    "includes_decor", "extra_child_price_usd",
                                    "extra_adult_price_usd", "cleaning_fee_usd",
                                    "features", "terms"])
    rows = fetch_all(cur, "party_packages", "WHERE is_active IS NOT FALSE",
                     "ORDER BY price_usd")
    for r in rows:
        feats_raw = r.get("features") or []
        terms_raw = r.get("additional_terms") or []
        features = [f for f in feats_raw if isinstance(f, str)]
        terms = []
        for t in terms_raw:
            if isinstance(t, dict):
                title = (t.get("title") or "").strip()
                desc = (t.get("description") or "").strip()
                terms.append(f"{title}: {desc}" if title else desc)

        kb.row("party_packages",
               package_id=r["package_id"], name=r.get("name"),
               price_usd=jsonable(r.get("price_usd")),
               base_children=r.get("base_children"),
               base_room_hours=jsonable(r.get("base_room_hours")),
               includes_food=bool(r.get("includes_food")),
               includes_drinks=bool(r.get("includes_drinks")),
               includes_decor=bool(r.get("includes_decor")),
               extra_child_price_usd=jsonable(r.get("extra_child_price")),
               extra_adult_price_usd=jsonable(r.get("extra_adult_price")),
               cleaning_fee_usd=jsonable(r.get("cleaning_fee")),
               features=" | ".join(features),
               terms=" | ".join(terms))

        name = r.get("name")
        price = money(r.get("price_usd"))
        kids = r.get("base_children") or 0
        hours = r.get("base_room_hours")
        hours_str = f"{hours:g}".rstrip("0").rstrip(".") if hours is not None else "—"

        included = []
        if r.get("includes_food"): included.append("food")
        if r.get("includes_drinks"): included.append("drinks")
        if r.get("includes_decor"): included.append("decorations")
        included_phrase = " It includes " + ", ".join(included) + "." if included else ""

        summary = (
            f"The {name} party package is {price} for up to {kids} children with "
            f"{hours_str} hour{'s' if hours_str != '1' else ''} of private party room time."
            f"{included_phrase}"
        )
        ec = r.get("extra_child_price")
        ea = r.get("extra_adult_price")
        if ec is not None or ea is not None:
            extras = []
            if ec is not None: extras.append(f"extra children {money(ec)} each")
            if ea is not None: extras.append(f"extra adults {money(ea)} each")
            summary += " " + "; ".join(extras).capitalize() + "."
        if features:
            summary += " Features: " + "; ".join(features) + "."

        doc_id = kb.add_doc(
            category="party_package", source_table="party_packages", entity_id=r["package_id"],
            name=f"{name} party package", summary=summary,
            tags=["party", "birthday", "package", name.lower().replace(" ", "-")],
            updated_at=r.get("updated_at"),
        )
        kb.add_fact(doc_id, "party_package", f"How much is the {name} party package?", price)
        kb.add_fact(doc_id, "party_package", f"How many children does the {name} package cover?",
                    f"{kids} children, with extra children at {money(ec)} each" if ec else f"{kids} children")
        if terms:
            kb.add_fact(doc_id, "party_package", f"What are the terms for the {name} package?",
                        "; ".join(terms))


def build_party_addons(kb: KB, cur):
    kb.add_sheet("party_addons", ["add_on_id", "code", "label", "price_usd",
                                   "price_type", "description"])
    rows = fetch_all(cur, "party_add_ons", "WHERE is_active IS NOT FALSE",
                     "ORDER BY display_order, add_on_id")
    for r in rows:
        kb.row("party_addons",
               add_on_id=r["add_on_id"], code=r.get("code"), label=r.get("label"),
               price_usd=jsonable(r.get("price")), price_type=r.get("price_type"),
               description=r.get("description"))

        label = r.get("label")
        price = money(r.get("price"))
        pt = r.get("price_type") or "flat"
        unit = {"flat": "", "perChild": " per child", "duration": " per hour"}.get(pt, f" ({pt})")
        desc = (r.get("description") or "").strip()
        summary = f"Party add-on '{label}' costs {price}{unit}. {desc}".strip()

        doc_id = kb.add_doc(
            category="party_addon", source_table="party_add_ons", entity_id=r["add_on_id"],
            name=label, summary=summary,
            tags=["party", "addon", "extra", r.get("code") or ""],
            updated_at=r.get("updated_at"),
        )
        kb.add_fact(doc_id, "party_addon", f"How much is the {label} add-on?", f"{price}{unit}")


def build_events(kb: KB, cur):
    kb.add_sheet("events", ["event_id", "title", "start_date", "end_date", "description"])
    rows = fetch_all(cur, "events", "WHERE is_published IS NOT FALSE", "ORDER BY start_date")
    for r in rows:
        kb.row("events",
               event_id=r["event_id"], title=r.get("title"),
               start_date=jsonable(r.get("start_date")),
               end_date=jsonable(r.get("end_date")),
               description=r.get("description"))

        title = (r.get("title") or "").strip()
        start = fmt_date(r.get("start_date"))
        end = fmt_date(r.get("end_date"))
        when = start if start == end or not end else f"{start} – {end}"
        desc = (r.get("description") or "").strip()
        summary = f"Event '{title}'"
        if when: summary += f" on {when}"
        summary += f". {desc}".rstrip(". ") + "."

        doc_id = kb.add_doc(
            category="event", source_table="events", entity_id=r["event_id"],
            name=title, summary=summary,
            tags=["event", "activity"],
            updated_at=r.get("updated_at"),
        )
        if when:
            kb.add_fact(doc_id, "event", f"When is '{title}'?", when)


def build_promotions(kb: KB, cur):
    kb.add_sheet("promotions", ["kind", "ref_id", "label_or_code", "scope", "discount",
                                 "starts_at", "ends_at", "description"])

    prod_rows = fetch_all(cur, "product_promotions", "WHERE is_active IS NOT FALSE",
                          "ORDER BY product_type, product_id")
    plan_lookup = {p["plan_id"]: p["name"]
                   for p in fetch_all(cur, "membership_plans", "", "ORDER BY plan_id")}

    for r in prod_rows:
        ptype = r.get("product_type")
        pid = r.get("product_id")
        target = f"{plan_lookup.get(pid, '?')} ({ptype} #{pid})" if ptype == "membership" else f"{ptype} #{pid}"
        disc_type = r.get("discount_type")
        dv = float(r.get("discount_value") or 0)
        discount = f"{dv:g}% off" if disc_type == "percent" else f"{money(dv)} off"
        starts = fmt_date(r.get("starts_at"))
        ends = fmt_date(r.get("ends_at"))
        label = r.get("promo_label") or "Promotion"
        note = (r.get("promo_note") or "").strip()

        kb.row("promotions",
               kind="product", ref_id=r["promotion_id"], label_or_code=label, scope=target,
               discount=discount, starts_at=jsonable(r.get("starts_at")),
               ends_at=jsonable(r.get("ends_at")), description=note)

        summary = f"Promotion '{label}': {discount} on {target}"
        if starts and ends:
            summary += f", valid {starts} – {ends}"
        summary += f". {note}".rstrip(". ") + "."
        doc_id = kb.add_doc(
            category="promotion", source_table="product_promotions", entity_id=r["promotion_id"],
            name=f"{label} ({target})", summary=summary,
            tags=["promotion", "discount", ptype or ""],
            updated_at=r.get("updated_at"),
        )
        kb.add_fact(doc_id, "promotion", f"Is there a promo on {target}?",
                    f"{label} — {discount}" + (f", through {ends}" if ends else ""))

    code_rows = fetch_all(cur, "promotions", "WHERE is_active IS NOT FALSE", "ORDER BY code")
    for r in code_rows:
        code = r.get("code")
        po = float(r.get("percent_off") or 0)
        ao = r.get("amount_off_usd")
        applies = r.get("applies_to") or []
        scope = ", ".join(applies) if applies else "all"
        starts = fmt_date(r.get("valid_from"))
        ends = fmt_date(r.get("valid_to"))
        discount = f"{po:g}% off" if po else f"{money(ao)} off"
        min_purch = r.get("min_purchase_usd")
        desc = (r.get("description") or "").strip()

        kb.row("promotions",
               kind="code", ref_id=r["promotion_id"], label_or_code=code, scope=scope,
               discount=discount, starts_at=jsonable(r.get("valid_from")),
               ends_at=jsonable(r.get("valid_to")), description=desc)

        summary = f"Promo code '{code}': {discount} on {scope}"
        if ends: summary += f", valid through {ends}"
        if min_purch: summary += f", min purchase {money(min_purch)}"
        summary += f". {desc}".rstrip(". ") + "."
        doc_id = kb.add_doc(
            category="promotion", source_table="promotions", entity_id=r["promotion_id"],
            name=f"Promo code {code}", summary=summary,
            tags=["promotion", "code", "discount"] + applies,
        )
        kb.add_fact(doc_id, "promotion", f"What does code {code} do?",
                    f"{discount} on {scope}" + (f" through {ends}" if ends else ""))


def build_pricing_fees(kb: KB, cur):
    kb.add_sheet("pricing_fees", ["key", "value", "unit", "value_display",
                                  "description", "audience"])
    rows = fetch_all(cur, "pricing_config", "WHERE is_active IS NOT FALSE", "ORDER BY config_key")

    PERCENT_KEYS = {"deposit_percentage", "tax_rate", "sibling_discount_rate"}
    ADMIN_ONLY = {"tax_rate"}  # internal — skip from KB

    for r in rows:
        key = r["config_key"]
        val = float(r.get("config_value") or 0)
        is_pct = key in PERCENT_KEYS
        unit = "%" if is_pct else "USD"
        display = f"{val:g}%" if is_pct else money(val)
        desc = (r.get("description") or "").strip()
        audience = "internal" if key in ADMIN_ONLY else "customer"

        kb.row("pricing_fees", key=key, value=val, unit=unit, value_display=display,
               description=desc, audience=audience)

        if audience == "internal":
            continue

        readable = key.replace("_", " ")
        summary = f"Playfunia's {readable} is {display}. {desc}".strip()
        doc_id = kb.add_doc(
            category="pricing_fee", source_table="pricing_config", entity_id=r["config_id"],
            name=readable, summary=summary,
            tags=["pricing", "fee", key],
            updated_at=r.get("updated_at"),
        )
        kb.add_fact(doc_id, "pricing_fee", f"How much is the {readable}?", display)


def build_jobs(kb: KB, cur):
    kb.add_sheet("jobs", ["listing_id", "title", "department", "employment_type",
                          "location", "pay_range", "minimum_age", "schedule_notes",
                          "summary_excerpt"])
    rows = fetch_all(cur, "job_listings", "WHERE is_active IS NOT FALSE",
                     "ORDER BY display_order, listing_id")
    for r in rows:
        title = r.get("title")
        desc = (r.get("description") or "").strip()
        excerpt = desc[:200] + ("…" if len(desc) > 200 else "")

        kb.row("jobs",
               listing_id=r["listing_id"], title=title, department=r.get("department"),
               employment_type=r.get("employment_type"), location=r.get("location"),
               pay_range=r.get("pay_range"), minimum_age=r.get("minimum_age"),
               schedule_notes=r.get("schedule_notes"), summary_excerpt=excerpt)

        bits = []
        if r.get("employment_type"): bits.append(r["employment_type"])
        if r.get("location"): bits.append(f"in {r['location']}")
        if r.get("pay_range"): bits.append(f"pay {r['pay_range']}")
        if r.get("minimum_age"): bits.append(f"min age {r['minimum_age']}")
        summary = f"{title}" + (f" — {', '.join(bits)}." if bits else ".") + f" {desc}"
        summary = summary.strip()

        kb.add_doc(
            category="job", source_table="job_listings", entity_id=r["listing_id"],
            name=title, summary=summary,
            tags=["careers", "job", "hiring", (r.get("department") or "").lower()],
            updated_at=r.get("updated_at"),
        )


# ---------- workbook writer -----------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
HEADER_FONT = Font(bold=True)


def write_sheet(wb: Workbook, name: str, cols: list[str], rows: list[list]):
    ws = wb.create_sheet(name[:31])
    ws.append(cols)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for row in rows:
        ws.append([
            json.dumps(jsonable(v), ensure_ascii=False) if isinstance(v, (dict, list)) else v
            for v in row
        ])
    for i, col_letter in enumerate(ws.iter_cols(min_row=1, max_row=1)):
        col_letter[0].alignment = Alignment(wrap_text=False)
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(cols) + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        width = min(60, max(12, len(cols[col_idx - 1]) + 2))
        ws.column_dimensions[letter].width = width


def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = psycopg2.connect(cursor_factory=psycopg2.extras.RealDictCursor, **PG)
    kb = KB()
    with conn.cursor() as cur:
        build_locations(kb, cur)
        build_hours(kb, cur)
        build_tickets(kb, cur)
        build_memberships(kb, cur)
        build_party_packages(kb, cur)
        build_party_addons(kb, cur)
        build_events(kb, cur)
        build_promotions(kb, cur)
        build_pricing_fees(kb, cur)
        build_jobs(kb, cur)
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    # knowledge_base first
    kb_cols = ["doc_id", "category", "slug", "name", "summary", "tags",
               "source_table", "entity_id", "source_updated_at"]
    write_sheet(wb, "knowledge_base", kb_cols,
                [[d[c] for c in kb_cols] for d in kb.docs])

    # facts
    write_sheet(wb, "facts", ["doc_id", "category", "question", "answer"],
                [[f["doc_id"], f["category"], f["question"], f["answer"]] for f in kb.facts])

    # structured sheets in a fixed order for predictability
    order = ["locations", "hours", "tickets", "memberships", "party_packages",
             "party_addons", "events", "promotions", "pricing_fees", "jobs"]
    for sn in order:
        if sn in kb.sheets:
            s = kb.sheets[sn]
            write_sheet(wb, sn, s["cols"], s["rows"])

    wb.save(OUT_PATH)

    print(f"Wrote {OUT_PATH}")
    print(f"  knowledge_base: {len(kb.docs)} docs")
    print(f"  facts:          {len(kb.facts)} Q&A pairs")
    for sn in order:
        if sn in kb.sheets:
            print(f"  {sn:<16} {len(kb.sheets[sn]['rows'])} rows")


if __name__ == "__main__":
    main()
